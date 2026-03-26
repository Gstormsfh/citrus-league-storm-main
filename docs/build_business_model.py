#!/usr/bin/env python3
"""
Citrus Fantasy Sports — Dynamic 24-Month Business Model
ALL computed cells use Excel formulas referencing Assumptions inputs.
Change any yellow input cell -> entire workbook recalculates instantly.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# Brand Colors
GREEN = "4A6741"
ORANGE = "D4702A"
CREAM = "F5F0E8"
LIGHT_GREEN = "D4E8B8"
INPUT_YELLOW = "FFF9E6"
RED = "CC0000"
WHITE = "FFFFFF"
LIGHT_ORANGE = "FFE8CC"

# Reusable Styles
header_font = Font(bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=GREEN)
input_fill = PatternFill("solid", fgColor=INPUT_YELLOW)
input_font = Font(bold=True, color=ORANGE)
total_fill = PatternFill("solid", fgColor=CREAM)
total_border_top = Border(top=Side(style="double"))
neg_font = Font(color=RED)
pos_font = Font(color=GREEN)
section_font = Font(bold=True, size=11, color=GREEN)
subsection_font = Font(bold=True, size=10)
in_season_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
off_season_fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
pct_fmt = "0.0%"
usd_fmt = "#,##0"
usd_cents_fmt = "#,##0.00"
num_fmt = "#,##0"

MONTHS = [
    "Apr-26","May-26","Jun-26","Jul-26","Aug-26","Sep-26",
    "Oct-26","Nov-26","Dec-26","Jan-27","Feb-27","Mar-27",
    "Apr-27","May-27","Jun-27","Jul-27","Aug-27","Sep-27",
    "Oct-27","Nov-27","Dec-27","Jan-28","Feb-28","Mar-28",
]
IS_IN_SEASON = [
    False, False, False, False, False, False,
    True,  True,  True,  True,  True,  True,
    False, False, False, False, False, False,
    True,  True,  True,  True,  True,  True,
]
# Growth story: 200 waitlist → June launch (50% invite 12 leaguemates) → draft explosion
NEW_USERS_DEFAULT = [
    100, 100, 1500, 3000, 6000, 10000,
    8000, 5000, 3000, 2500, 2000, 1500,
    1500, 2000, 3000, 5000, 8000, 12000,
    17000, 20000, 20000, 18000, 16000, 13000,
]
# Content marketing as lump-sum influencer payments per month
CONTENT_MKT_DEFAULT = [
    0, 0, 500, 0, 0, 1500,
    0, 0, 0, 0, 0, 0,
    0, 0, 2000, 0, 2500, 3000,
    3000, 0, 2000, 0, 0, 0,
]
CONFERENCE_COSTS = {
    0: 750, 3: 4500, 5: 3500, 7: 5000, 10: 4000, 11: 5000,
    12: 500, 14: 6000, 15: 6000, 17: 5000, 19: 8000, 22: 5000, 23: 6000,
}
# US states with DFS licensing — (state_key, label) for formula building
US_DFS_STATES = [
    "ny", "ma", "va", "tn", "in", "pa", "ct", "co", "il", "mo",
    "or", "md", "ks", "vt", "ar", "ms", "nh", "de", "ri",
]

# Assumption layout: (ref_key, label, value, unit, source, cell_format)
# ref_key=None for section headers / blanks (not referenced by formulas)
ASSUMPTIONS = [
    (None, "PRICING", None, None, None, None),
    ("pro_price", "Pro Tier \u2014 Monthly", 4.99, "USD/mo", "", usd_cents_fmt),
    ("comm_price", "Commissioner Tier \u2014 Monthly", 9.99, "USD/mo", "", usd_cents_fmt),
    ("annual_disc", "Annual Billing Discount", 0.20, "%", "Industry standard", pct_fmt),
    ("annual_pct", "Annual Billing %", 0.30, "%", "Industry benchmarks", pct_fmt),
    ("monthly_pct", "Monthly Billing %", 0.70, "%", "", pct_fmt),
    (None, "", None, None, None, None),
    (None, "CONVERSION RATES", None, None, None, None),
    ("premium_thresh", "Premium Features Launch Threshold", 10000, "users", "No paid tiers until user base hits this", num_fmt),
    ("conv_pro_y1", "Free \u2192 Pro Y1", 0.025, "%", "RevenueCat median 2.18%; niche premium", pct_fmt),
    ("conv_pro_y2", "Free \u2192 Pro Y2", 0.055, "%", "Product maturation + mobile", pct_fmt),
    ("conv_comm_y1", "Free \u2192 Commissioner Y1", 0.008, "%", "Power-user tier", pct_fmt),
    ("conv_comm_y2", "Free \u2192 Commissioner Y2", 0.015, "%", "", pct_fmt),
    ("mult_in", "In-Season Multiplier", 1.8, "x", "NHL engagement seasonality", '0.0"x"'),
    ("mult_off", "Off-Season Multiplier", 0.4, "x", "", '0.0"x"'),
    (None, "", None, None, None, None),
    (None, "CHURN RATES (Monthly)", None, None, None, None),
    ("ch_pro_in", "Pro In-Season", 0.04, "%", "Below mobile avg 9%", pct_fmt),
    ("ch_pro_off", "Pro Off-Season", 0.12, "%", "Seasonal disengagement", pct_fmt),
    ("ch_comm_in", "Commissioner In-Season", 0.025, "%", "Stickier \u2014 league mgmt", pct_fmt),
    ("ch_comm_off", "Commissioner Off-Season", 0.08, "%", "", pct_fmt),
    ("ch_free_in", "Free In-Season", 0.06, "%", "", pct_fmt),
    ("ch_free_off", "Free Off-Season", 0.15, "%", "", pct_fmt),
    (None, "", None, None, None, None),
    (None, "STORMY AI — TOKEN BUDGETS (per tier)", None, None, None, None),
    ("tok_free", "Free Tier Tokens/mo", 5000, "tokens", "~2-3 short questions", num_fmt),
    ("tok_pro", "Pro Tier Tokens/mo", 50000, "tokens", "~25 questions/mo", num_fmt),
    ("tok_comm", "Commissioner Tier Tokens/mo", 150000, "tokens", "~75 questions/mo", num_fmt),
    (None, "", None, None, None, None),
    (None, "PAY-PER-USE (Token Top-Up Packs)", None, None, None, None),
    ("topup_price_free", "Free User Token Pack", 2.99, "USD", "25K tokens \u2014 $0.12/1K (premium)", usd_cents_fmt),
    ("topup_price_paid", "Paid User Token Pack", 1.99, "USD", "100K tokens \u2014 $0.02/1K (subscriber perk)", usd_cents_fmt),
    ("topup_attach_free", "Top-Up Attach \u2014 Free Users", 0.03, "%/mo", "Hit limit, want more", pct_fmt),
    ("topup_attach_paid", "Top-Up Attach \u2014 Pro/Comm", 0.08, "%/mo", "Power users exceeding budget", pct_fmt),
    ("dk_price", "Draft Kit Price", 4.99, "USD", "One-time per season", usd_cents_fmt),
    ("dk_attach", "Draft Kit Attach Rate", 0.08, "%", "All users in draft months", pct_fmt),
    (None, "", None, None, None, None),
    (None, "DFS", None, None, None, None),
    ("dfs_rake", "DFS Rake", 0.12, "%", "DK/FD range 10-16%", pct_fmt),
    ("dfs_entry", "Average Entry Fee", 5.00, "USD", "", usd_cents_fmt),
    ("dfs_part", "Participation Rate", 0.05, "%", "Hockey-only; DK ~10% all sports", pct_fmt),
    ("dfs_entries", "Entries per User per Month", 5, "#", "", num_fmt),
    ("dfs_launch", "DFS Launch Month Index", 18, "0-based", "Oct 2027", num_fmt),
    (None, "", None, None, None, None),
    (None, "ADVERTISING (web only — no mobile app ads)", None, None, None, None),
    ("ecpm", "Blended eCPM", 6.00, "USD", "BusinessOfApps/Appodeal — web display", usd_cents_fmt),
    ("pv_in", "Pageviews/Free User In-Season", 100, "#/mo", "", num_fmt),
    ("pv_off", "Pageviews/Free User Off-Season", 50, "#/mo", "", num_fmt),
    ("ad_elig", "Ad-Eligible % (web users only)", 0.24, "%", "30% web × 80% no ad-blocker", pct_fmt),
    (None, "", None, None, None, None),
    (None, "INFRASTRUCTURE", None, None, None, None),
    ("sb_pro", "Supabase Pro", 25, "USD/mo", "supabase.com/pricing", usd_fmt),
    ("sb_team", "Supabase Team (>threshold)", 599, "USD/mo", "supabase.com/pricing", usd_fmt),
    ("sb_thresh", "Supabase User Threshold", 50000, "users", "", num_fmt),
    ("fb_base", "Firebase Base Cost", 5, "USD/mo", "firebase.google.com", usd_fmt),
    ("fb_scale", "Firebase Scaling per User", 0.001, "USD/user", "Rough estimate", "0.0000"),
    ("cl_cost_per_m", "Claude API Cost per 1M Tokens", 5.00, "USD/1M tok", "Blended Sonnet: $3 in + $15 out + caching", usd_cents_fmt),
    ("proxy", "Proxy Rotation", 100, "USD/mo", "", usd_fmt),
    ("domain", "Domain / SSL / Misc", 20, "USD/mo", "", usd_fmt),
    (None, "", None, None, None, None),
    (None, "PLATFORM FEES", None, None, None, None),
    ("apple_comm", "Apple App Store Commission", 0.15, "%", "Small Business Program", pct_fmt),
    ("google_comm", "Google Play Commission", 0.15, "%", "Reduced service fee", pct_fmt),
    ("stripe_pct", "Stripe Processing %", 0.029, "%", "stripe.com/pricing", "0.00%"),
    ("stripe_flat", "Stripe Flat Fee", 0.30, "USD/txn", "", usd_cents_fmt),
    ("ios_split", "Mobile Revenue \u2014 iOS %", 0.60, "%", "", pct_fmt),
    ("android_split", "Mobile Revenue \u2014 Android %", 0.40, "%", "", pct_fmt),
    ("mobile_pct", "Mobile vs Web \u2014 Mobile %", 0.70, "%", "", pct_fmt),
    ("apple_dev", "Apple Developer Annual", 99, "USD/yr", "", usd_fmt),
    ("google_reg", "Google Play Registration", 25, "USD", "One-time", usd_fmt),
    (None, "", None, None, None, None),
    (None, "MARKETING", None, None, None, None),
    ("mkt_launch", "Marketing Launch Month", 2, "0-based", "Mid-June 2026 app launch", num_fmt),
    ("cac", "Blended CAC Target", 2.50, "USD", "Organic/community-driven", usd_cents_fmt),
    (None, "", None, None, None, None),
    (None, "CONTENT / INFLUENCER MARKETING (Monthly Lump Sums)", None, None, None, None),
] + [
    (f"cont_mkt_{i}", f"  Influencer Spend \u2014 {MONTHS[i]}", CONTENT_MKT_DEFAULT[i], "USD", "Lump-sum influencer payment", usd_fmt)
    for i in range(24)
] + [
    (None, "", None, None, None, None),
    (None, "TEAM", None, None, None, None),
    ("team_1_6", "Months 1-6 (Solo Founder)", 0, "USD/mo", "Bootstrapped", usd_fmt),
    ("team_7_12", "Months 7-12 (1 PT Contractor)", 1500, "USD/mo", "", usd_fmt),
    ("team_13_18", "Months 13-18 (1-2 Contractors)", 3000, "USD/mo", "", usd_fmt),
    ("team_19_24", "Months 19-24 (Small Team)", 6000, "USD/mo", "", usd_fmt),
    (None, "", None, None, None, None),
    (None, "LEGAL / ACCOUNTING", None, None, None, None),
    ("legal_y1", "Year 1 (Basic)", 100, "USD/mo", "Bookkeeping, tax prep only", usd_fmt),
    ("legal_y2", "Year 2 (Basic)", 300, "USD/mo", "Bookkeeping + general counsel", usd_fmt),
    (None, "", None, None, None, None),
    (None, "DFS COMPLIANCE — CANADA (NO LICENSE REQUIRED)", None, None, None, None),
    ("dfs_reg_thresh", "Compliance Threshold", 100000, "users", "Prudent compliance at scale", num_fmt),
    ("dfs_legal_opinion", "Legal Opinion / Skill-Game Defense", 500, "USD/mo", "s.206(1)(d) opinion amortized; NO license needed", usd_fmt),
    ("dfs_compliance", "FINTRAC AML Compliance", 1000, "USD/mo", "Prize pools >$10K trigger reporting", usd_fmt),
    ("dfs_insurance", "Insurance (E&O + Cyber)", 500, "USD/mo", "Standard biz insurance", usd_fmt),
    ("dfs_player_trust", "Player Fund Segregation / Audit", 500, "USD/mo", "Trust account — best practice, not required", usd_fmt),
    (None, "", None, None, None, None),
    (None, "US DFS EXPANSION (Toggle)", None, None, None, None),
    ("us_expand_month", "US Expansion Launch Month", 99, "0-based", "Set to 99 = OFF; e.g. 18 = Oct-27", num_fmt),
    (None, "", None, None, None, None),
    (None, "  STATE-BY-STATE LICENSING FEES (Year 1 / Annual)", None, None, None, None),
    ("us_ny_y1", "  New York — Year 1", 500000, "USD", "NY Gaming Commission — most expensive state", usd_fmt),
    ("us_ny_ann", "  New York — Annual Renewal", 50000, "USD/yr", "NY Gaming Commission", usd_fmt),
    ("us_ma_y1", "  Massachusetts — Year 1", 50000, "USD", "MA Gaming Commission", usd_fmt),
    ("us_ma_ann", "  Massachusetts — Annual Renewal", 50000, "USD/yr", "MA Gaming Commission", usd_fmt),
    ("us_va_y1", "  Virginia — Year 1", 50000, "USD", "VA Lottery", usd_fmt),
    ("us_va_ann", "  Virginia — Annual Renewal", 50000, "USD/yr", "VA Lottery", usd_fmt),
    ("us_tn_y1", "  Tennessee — Year 1", 50000, "USD", "TN Secretary of State", usd_fmt),
    ("us_tn_ann", "  Tennessee — Annual Renewal", 50000, "USD/yr", "TN Secretary of State", usd_fmt),
    ("us_in_y1", "  Indiana — Year 1", 50000, "USD", "IN Gaming Commission", usd_fmt),
    ("us_in_ann", "  Indiana — Annual Renewal", 5000, "USD/yr", "IN Gaming Commission", usd_fmt),
    ("us_pa_y1", "  Pennsylvania — Year 1", 50000, "USD", "PA Gaming Control Board", usd_fmt),
    ("us_pa_ann", "  Pennsylvania — Annual Renewal", 10000, "USD/yr", "PA Gaming Control Board", usd_fmt),
    ("us_ct_y1", "  Connecticut — Year 1", 15000, "USD", "CT Dept Consumer Protection", usd_fmt),
    ("us_ct_ann", "  Connecticut — Annual Renewal", 15000, "USD/yr", "CT Dept Consumer Protection", usd_fmt),
    ("us_co_y1", "  Colorado — Year 1", 10000, "USD", "CO Dept of Revenue", usd_fmt),
    ("us_co_ann", "  Colorado — Annual Renewal", 10000, "USD/yr", "CO Dept of Revenue", usd_fmt),
    ("us_il_y1", "  Illinois — Year 1", 10000, "USD", "IL Gaming Board", usd_fmt),
    ("us_il_ann", "  Illinois — Annual Renewal", 10000, "USD/yr", "IL Gaming Board", usd_fmt),
    ("us_mo_y1", "  Missouri — Year 1", 10000, "USD", "MO Gaming Commission", usd_fmt),
    ("us_mo_ann", "  Missouri — Annual Renewal", 10000, "USD/yr", "MO Gaming Commission", usd_fmt),
    ("us_or_y1", "  Oregon — Year 1", 10000, "USD", "OR Dept of Justice", usd_fmt),
    ("us_or_ann", "  Oregon — Annual Renewal", 10000, "USD/yr", "OR Dept of Justice", usd_fmt),
    ("us_md_y1", "  Maryland — Year 1", 5000, "USD", "MD Lottery & Gaming", usd_fmt),
    ("us_md_ann", "  Maryland — Annual Renewal", 5000, "USD/yr", "MD Lottery & Gaming", usd_fmt),
    ("us_ks_y1", "  Kansas — Year 1", 5000, "USD", "KS Attorney General", usd_fmt),
    ("us_ks_ann", "  Kansas — Annual Renewal", 5000, "USD/yr", "KS Attorney General", usd_fmt),
    ("us_vt_y1", "  Vermont — Year 1", 5000, "USD", "VT Attorney General", usd_fmt),
    ("us_vt_ann", "  Vermont — Annual Renewal", 5000, "USD/yr", "VT Attorney General", usd_fmt),
    ("us_ar_y1", "  Arkansas — Year 1", 5000, "USD", "AR Dept Finance & Admin", usd_fmt),
    ("us_ar_ann", "  Arkansas — Annual Renewal", 5000, "USD/yr", "AR Dept Finance & Admin", usd_fmt),
    ("us_ms_y1", "  Mississippi — Year 1", 5000, "USD", "MS Gaming Commission", usd_fmt),
    ("us_ms_ann", "  Mississippi — Annual Renewal", 5000, "USD/yr", "MS Gaming Commission", usd_fmt),
    ("us_nh_y1", "  New Hampshire — Year 1", 5000, "USD", "NH Lottery Commission", usd_fmt),
    ("us_nh_ann", "  New Hampshire — Annual Renewal", 5000, "USD/yr", "NH Lottery Commission", usd_fmt),
    ("us_de_y1", "  Delaware — Year 1", 5000, "USD", "DE Dept of Finance", usd_fmt),
    ("us_de_ann", "  Delaware — Annual Renewal", 5000, "USD/yr", "DE Dept of Finance", usd_fmt),
    ("us_ri_y1", "  Rhode Island — Year 1", 5000, "USD", "RI DBR", usd_fmt),
    ("us_ri_ann", "  Rhode Island — Annual Renewal", 5000, "USD/yr", "RI DBR", usd_fmt),
    (None, "", None, None, None, None),
    (None, "  US COMPLIANCE OVERHEAD (Annual)", None, None, None, None),
    ("us_legal_counsel", "  Multi-State Legal Counsel", 150000, "USD/yr", "DFS compliance firm", usd_fmt),
    ("us_geolocation", "  Geolocation Vendor (GeoComply etc.)", 50000, "USD/yr", "Block banned states, verify regulated", usd_fmt),
    ("us_audit", "  Independent Financial Audit", 35000, "USD/yr", "Required by most regulated states", usd_fmt),
    ("us_responsible_gaming", "  Responsible Gaming Program", 20000, "USD/yr", "Self-exclusion, limits, reporting", usd_fmt),
    ("us_reporting", "  State Reporting & Filings", 20000, "USD/yr", "Annual filings, revenue reports", usd_fmt),
    ("us_bg_checks", "  Background Checks (Key Personnel)", 10000, "USD/yr", "Required for all licensed states", usd_fmt),
    ("us_ny_rev_tax", "  NY Revenue Tax Rate", 0.15, "%", "~15% of DFS revenue from NY users", pct_fmt),
    ("us_ny_rev_share", "  Est. NY Revenue Share (% of total)", 0.12, "%", "NY ~12% of US DFS market", pct_fmt),
    (None, "", None, None, None, None),
    (None, "USER GROWTH / ACQUISITION (Monthly New Users)", None, None, None, None),
] + [
    (f"new_users_{i}", f"  New Users — {MONTHS[i]}", NEW_USERS_DEFAULT[i], "users", "Monthly new user target", num_fmt)
    for i in range(24)
] + [
    (None, "", None, None, None, None),
    (None, "CASH", None, None, None, None),
    ("start_cash", "Starting Cash", 1000, "USD", "", usd_fmt),
]

# Market reference data (informational only, not linked to formulas)
MARKET_REFERENCE = [
    (None, "MARKET REFERENCE (informational)", None, None, None, None),
    (None, "TAM \u2014 Fantasy Sports (2025)", "$32B+", "USD", "Mordor Intelligence", None),
    (None, "SAM \u2014 Fantasy Hockey Global", "$1-2B", "USD", "Industry estimates", None),
    (None, "Active Fantasy Hockey Players (NA)", "5,800,000", "Users", "FSGA industry data", None),
    (None, "NHL Fantasy CAGR", "14.6%", "%", "Mordor Intelligence", None),
    (None, "", None, None, None, None),
    (None, "BENCHMARKS", None, None, None, None),
    (None, "Mobile Freemium Median", "2.18%", "%", "RevenueCat 2025", None),
    (None, "Sleeper Annual Retention", "75%", "%", "Sleeper Series C PR", None),
    (None, "DraftKings/FanDuel CAC", "$200-350", "USD", "BusinessOfApps", None),
    (None, "NA Mobile CPI", "$5.28", "USD", "BusinessOfApps", None),
]

# ── User Growth row assignments ──
UG_SEASON = 2
UG_NEW = 3
UG_EFF_CONV_PRO = 5
UG_EFF_CONV_COMM = 6
UG_CH_FREE_RATE = 7
UG_CH_PRO_RATE = 8
UG_CH_COMM_RATE = 9
UG_CHURNED_FREE = 11
UG_CHURNED_PRO = 12
UG_CHURNED_COMM = 13
UG_TO_PRO = 14
UG_TO_COMM = 15
UG_END_FREE = 17
UG_END_PRO = 18
UG_END_COMM = 19
UG_END_TOTAL = 20
UG_PREMIUM_PCT = 22
UG_BLENDED_CHURN = 23
UG_RETENTION = 24

# ── Revenue row assignments ──
R_PRO_MO = 3
R_PRO_AN = 4
R_COMM_MO = 5
R_COMM_AN = 6
R_GROSS_SUB = 7
R_DISC = 8
R_NET_SUB = 9
R_AI_FREE = 12
R_AI_PRO = 13
R_DRAFT_KIT = 14
R_NET_PPU = 15
R_DFS_GROSS = 18
R_DFS_RAKE = 19
R_NET_DFS = 20
R_AD_IMPS = 23
R_AD_REV = 24
R_NET_AD = 25
R_GROSS_REV = 27
R_APPLE_FEE = 30
R_GOOGLE_FEE = 31
R_STRIPE_FEE = 32
R_TOTAL_PF = 33
R_NET_REV = 35
R_SUB_PCT = 38
R_PPU_PCT = 39
R_DFS_PCT = 40
R_AD_PCT = 41
R_MRR = 43
R_ARR = 44
R_CUM_REV = 45

# ── COGS & OPEX row assignments ──
CE_SUPABASE = 3
CE_FIREBASE = 4
CE_CLAUDE = 5
CE_PROXY = 6
CE_DOMAIN = 7
CE_TOTAL_COGS = 8
CE_GROSS_PROFIT = 10
CE_GROSS_MARGIN = 11
CE_DIG_MKT = 15
CE_CONT_MKT = 16
CE_TOTAL_SM = 17
CE_TEAM = 20
CE_LEGAL = 21
CE_DFS_REG = 22
CE_US_DFS_LIC = 23
CE_US_DFS_COMPLY = 24
CE_US_DFS_TOTAL = 25
CE_APPLE_DEV = 26
CE_GOOGLE_REG = 27
CE_TOTAL_GA = 28
CE_CONF = 31
CE_TOTAL_CONF = 32
CE_TOTAL_OPEX = 34
CE_EBITDA = 36
CE_EBITDA_MARGIN = 37
CE_NET_INCOME = 39

# ── Unit Economics row assignments ──
UE_ARPU_TOTAL = 2
UE_ARPU_PAID = 3
UE_PAID_SUBS = 4
UE_BLEND_CHURN = 5
UE_LTV = 6
UE_CAC = 7
UE_LTV_CAC = 8
UE_PAYBACK = 9
UE_TEAM_SIZE = 10
UE_REV_EMP = 11

# ── Cash Flow row assignments ──
CF_START = 2
CF_MONTHLY = 3
CF_CUM = 4
CF_BURN = 5
CF_RUNWAY = 6


# ── Helpers ─────────────────────────────────────────────────────────────

def write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_total_row(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        ws.cell(row=row, column=col).fill = total_fill
        ws.cell(row=row, column=col).border = total_border_top


def c(col):
    """Column number to letter."""
    return get_column_letter(col)


def ug(row, col):
    """Reference a cell on the User Growth sheet."""
    return f"'User Growth'!{c(col)}{row}"


def rev(row, col):
    """Reference a cell on the Revenue sheet."""
    return f"Revenue!{c(col)}{row}"


def cogs(row, col):
    """Reference a cell on the COGS & Operating Expenses sheet."""
    return f"'COGS & Operating Expenses'!{c(col)}{row}"


def style_formula_row(ws, row, start_col, end_col, fmt, is_total=False):
    """Apply number format and alignment to a formula row."""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right")
    if is_total:
        write_total_row(ws, row, 1, end_col)


def write_section(ws, row, label):
    ws.cell(row=row, column=1, value=label).font = section_font


def write_label(ws, row, label, bold=False):
    cell = ws.cell(row=row, column=1, value=label)
    if bold:
        cell.font = subsection_font
    else:
        cell.font = Font(size=10)



# ══════════════════════════════════════════════════════════════════════════
# WORKBOOK BUILDER
# ══════════════════════════════════════════════════════════════════════════

def build_workbook():
    wb = Workbook()

    # ──────────────────────────────────────────────────────────────────
    # TAB 1: ASSUMPTIONS — Editable numeric inputs
    # ──────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Assumptions"
    set_col_widths(ws, [40, 18, 14, 65])
    write_header_row(ws, 1, ["Assumption", "Value", "Unit", "Source / Citation"])

    A = {}  # ref_key -> "Assumptions!$B$<row>"
    row = 2
    for item in ASSUMPTIONS:
        ref_key, label, value, unit, source, fmt = item
        ws.cell(row=row, column=1, value=label)
        if label and label.isupper() and value is None:
            ws.cell(row=row, column=1).font = section_font
        elif value is not None:
            ws.cell(row=row, column=1).font = Font(size=10)
            cell = ws.cell(row=row, column=2, value=value)
            cell.fill = input_fill
            cell.font = input_font
            cell.alignment = Alignment(horizontal="center")
            if fmt:
                cell.number_format = fmt
            ws.cell(row=row, column=3, value=unit or "").alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4, value=source or "")
        if ref_key:
            A[ref_key] = f"Assumptions!$B${row}"
        row += 1

    # Add market reference data (informational)
    row += 1
    for item in MARKET_REFERENCE:
        _, label, value, unit, source, _ = item
        ws.cell(row=row, column=1, value=label)
        if label and label.isupper() and value is None:
            ws.cell(row=row, column=1).font = section_font
        elif value is not None:
            ws.cell(row=row, column=1).font = Font(size=10)
            ws.cell(row=row, column=2, value=value).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=unit or "").alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4, value=source or "")
        row += 1

    row += 1
    ws.cell(row=row, column=1,
            value="Edit any yellow cell above to recalculate the entire model.").font = Font(italic=True, color=GREEN)

    # ──────────────────────────────────────────────────────────────────
    # TAB 2: USER GROWTH — All formulas
    # ──────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("User Growth")
    set_col_widths(ws2, [32] + [13] * 24)
    write_header_row(ws2, 1, ["Metric"] + MONTHS)

    # Season status row
    write_label(ws2, UG_SEASON, "NHL Season Status", bold=True)
    for m in range(24):
        col = m + 2
        cell = ws2.cell(row=UG_SEASON, column=col)
        if IS_IN_SEASON[m]:
            cell.value = "IN-SEASON"
            cell.fill = in_season_fill
        else:
            cell.value = "OFF-SEASON"
            cell.fill = off_season_fill
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(size=9)

    # New Users row (FORMULA — references Assumptions tab inputs)
    write_label(ws2, UG_NEW, "New Users Acquired", bold=True)
    for m in range(24):
        col = m + 2
        cell = ws2.cell(row=UG_NEW, column=col, value=f"={A[f'new_users_{m}']}")
        cell.number_format = num_fmt
        cell.alignment = Alignment(horizontal="right")

    # Blank row 4
    # Effective rates section
    write_label(ws2, UG_EFF_CONV_PRO, "  Eff. Conv Rate → Pro")
    write_label(ws2, UG_EFF_CONV_COMM, "  Eff. Conv Rate → Comm")
    write_label(ws2, UG_CH_FREE_RATE, "  Churn Rate (Free)")
    write_label(ws2, UG_CH_PRO_RATE, "  Churn Rate (Pro)")
    write_label(ws2, UG_CH_COMM_RATE, "  Churn Rate (Comm)")
    write_label(ws2, UG_CHURNED_FREE, "Churned Free Users")
    write_label(ws2, UG_CHURNED_PRO, "Churned Pro Users")
    write_label(ws2, UG_CHURNED_COMM, "Churned Commissioner Users")
    write_label(ws2, UG_TO_PRO, "Conversions to Pro")
    write_label(ws2, UG_TO_COMM, "Conversions to Commissioner")
    write_label(ws2, UG_END_FREE, "End-of-Month Free Users")
    write_label(ws2, UG_END_PRO, "End-of-Month Pro Users")
    write_label(ws2, UG_END_COMM, "End-of-Month Commissioner Users")
    write_label(ws2, UG_END_TOTAL, "End-of-Month TOTAL Users", bold=True)
    write_label(ws2, UG_PREMIUM_PCT, "Premium % of Total")
    write_label(ws2, UG_BLENDED_CHURN, "Blended Monthly Churn Rate")
    write_label(ws2, UG_RETENTION, "Monthly Retention Rate")

    for m in range(24):
        col = m + 2
        cl = c(col)
        pcl = c(col - 1) if m > 0 else None
        prev_free = f"{pcl}{UG_END_FREE}" if m > 0 else "0"
        prev_pro = f"{pcl}{UG_END_PRO}" if m > 0 else "0"
        prev_comm = f"{pcl}{UG_END_COMM}" if m > 0 else "0"

        # Effective conversion rate to Pro (0 until user base hits premium threshold)
        prev_total = f"{pcl}{UG_END_TOTAL}" if m > 0 else "0"
        base_conv_pro = f"IF(COLUMN()<14,{A['conv_pro_y1']},{A['conv_pro_y2']})*IF({cl}{UG_SEASON}=\"IN-SEASON\",{A['mult_in']},{A['mult_off']})"
        ws2.cell(row=UG_EFF_CONV_PRO, column=col,
                 value=f"=IF({prev_total}>={A['premium_thresh']},{base_conv_pro},0)")
        # Effective conversion rate to Commissioner (same gate)
        base_conv_comm = f"IF(COLUMN()<14,{A['conv_comm_y1']},{A['conv_comm_y2']})*IF({cl}{UG_SEASON}=\"IN-SEASON\",{A['mult_in']},{A['mult_off']})"
        ws2.cell(row=UG_EFF_CONV_COMM, column=col,
                 value=f"=IF({prev_total}>={A['premium_thresh']},{base_conv_comm},0)")
        # Churn rates
        ws2.cell(row=UG_CH_FREE_RATE, column=col,
                 value=f"=IF({cl}{UG_SEASON}=\"IN-SEASON\",{A['ch_free_in']},{A['ch_free_off']})")
        ws2.cell(row=UG_CH_PRO_RATE, column=col,
                 value=f"=IF({cl}{UG_SEASON}=\"IN-SEASON\",{A['ch_pro_in']},{A['ch_pro_off']})")
        ws2.cell(row=UG_CH_COMM_RATE, column=col,
                 value=f"=IF({cl}{UG_SEASON}=\"IN-SEASON\",{A['ch_comm_in']},{A['ch_comm_off']})")

        # Churned users = INT(previous_users * churn_rate)
        ws2.cell(row=UG_CHURNED_FREE, column=col,
                 value=f"=INT({prev_free}*{cl}{UG_CH_FREE_RATE})")
        ws2.cell(row=UG_CHURNED_PRO, column=col,
                 value=f"=INT({prev_pro}*{cl}{UG_CH_PRO_RATE})")
        ws2.cell(row=UG_CHURNED_COMM, column=col,
                 value=f"=INT({prev_comm}*{cl}{UG_CH_COMM_RATE})")

        # Conversions from free base
        ws2.cell(row=UG_TO_PRO, column=col,
                 value=f"=INT({prev_free}*{cl}{UG_EFF_CONV_PRO})")
        ws2.cell(row=UG_TO_COMM, column=col,
                 value=f"=INT({prev_free}*{cl}{UG_EFF_CONV_COMM})")

        # End-of-month pools
        ws2.cell(row=UG_END_FREE, column=col,
                 value=f"=MAX(0,{prev_free}+{cl}{UG_NEW}-{cl}{UG_CHURNED_FREE}-{cl}{UG_TO_PRO}-{cl}{UG_TO_COMM})")
        ws2.cell(row=UG_END_PRO, column=col,
                 value=f"=MAX(0,{prev_pro}+{cl}{UG_TO_PRO}-{cl}{UG_CHURNED_PRO})")
        ws2.cell(row=UG_END_COMM, column=col,
                 value=f"=MAX(0,{prev_comm}+{cl}{UG_TO_COMM}-{cl}{UG_CHURNED_COMM})")
        ws2.cell(row=UG_END_TOTAL, column=col,
                 value=f"={cl}{UG_END_FREE}+{cl}{UG_END_PRO}+{cl}{UG_END_COMM}")

        # Derived metrics
        ws2.cell(row=UG_PREMIUM_PCT, column=col,
                 value=f"=IF({cl}{UG_END_TOTAL}>0,({cl}{UG_END_PRO}+{cl}{UG_END_COMM})/{cl}{UG_END_TOTAL},0)")
        ws2.cell(row=UG_BLENDED_CHURN, column=col,
                 value=f"=IF({cl}{UG_END_TOTAL}+{cl}{UG_CHURNED_FREE}+{cl}{UG_CHURNED_PRO}+{cl}{UG_CHURNED_COMM}>0,({cl}{UG_CHURNED_FREE}+{cl}{UG_CHURNED_PRO}+{cl}{UG_CHURNED_COMM})/({cl}{UG_END_TOTAL}+{cl}{UG_CHURNED_FREE}+{cl}{UG_CHURNED_PRO}+{cl}{UG_CHURNED_COMM}),0)")
        ws2.cell(row=UG_RETENTION, column=col,
                 value=f"=1-{cl}{UG_BLENDED_CHURN}")

    # Apply formats
    for r in [UG_EFF_CONV_PRO, UG_EFF_CONV_COMM, UG_CH_FREE_RATE, UG_CH_PRO_RATE, UG_CH_COMM_RATE]:
        style_formula_row(ws2, r, 2, 25, pct_fmt)
    for r in [UG_CHURNED_FREE, UG_CHURNED_PRO, UG_CHURNED_COMM, UG_TO_PRO, UG_TO_COMM,
              UG_END_FREE, UG_END_PRO, UG_END_COMM, UG_END_TOTAL]:
        style_formula_row(ws2, r, 2, 25, num_fmt, is_total=(r == UG_END_TOTAL))
    for r in [UG_PREMIUM_PCT, UG_BLENDED_CHURN, UG_RETENTION]:
        style_formula_row(ws2, r, 2, 25, pct_fmt)


    # ──────────────────────────────────────────────────────────────────
    # TAB 3: REVENUE — All formulas
    # ──────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Revenue")
    set_col_widths(ws3, [38] + [14] * 24)
    write_header_row(ws3, 1, ["Revenue Line"] + MONTHS)

    # Labels
    write_section(ws3, 2, "SUBSCRIPTION REVENUE")
    write_label(ws3, R_PRO_MO, "  Pro Monthly Billing")
    write_label(ws3, R_PRO_AN, "  Pro Annual Billing")
    write_label(ws3, R_COMM_MO, "  Commissioner Monthly Billing")
    write_label(ws3, R_COMM_AN, "  Commissioner Annual Billing")
    write_label(ws3, R_GROSS_SUB, "  Gross Subscription Revenue")
    write_label(ws3, R_DISC, "  Less: Annual Billing Discount")
    write_label(ws3, R_NET_SUB, "  NET SUBSCRIPTION REVENUE", bold=True)
    write_section(ws3, 11, "PAY-PER-USE REVENUE")
    write_label(ws3, R_AI_FREE, "  Token Top-Ups (Free @ $2.99)")
    write_label(ws3, R_AI_PRO, "  Token Top-Ups (Paid @ $1.99)")
    write_label(ws3, R_DRAFT_KIT, "  Draft Kits (Seasonal)")
    write_label(ws3, R_NET_PPU, "  NET PAY-PER-USE REVENUE", bold=True)
    write_section(ws3, 17, "DFS REVENUE")
    write_label(ws3, R_DFS_GROSS, "  Gross DFS Entry Fees (info)")
    write_label(ws3, R_DFS_RAKE, "  DFS Rake Revenue")
    write_label(ws3, R_NET_DFS, "  NET DFS REVENUE", bold=True)
    write_section(ws3, 22, "ADVERTISING REVENUE")
    write_label(ws3, R_AD_IMPS, "  Ad Impressions (000s)")
    write_label(ws3, R_AD_REV, "  Ad Revenue")
    write_label(ws3, R_NET_AD, "  NET ADVERTISING REVENUE", bold=True)
    write_label(ws3, R_GROSS_REV, "GROSS REVENUE", bold=True)
    write_section(ws3, 29, "PLATFORM FEES")
    write_label(ws3, R_APPLE_FEE, "  Apple App Store (15%)")
    write_label(ws3, R_GOOGLE_FEE, "  Google Play (15%)")
    write_label(ws3, R_STRIPE_FEE, "  Stripe Processing")
    write_label(ws3, R_TOTAL_PF, "  Total Platform Fees")
    write_label(ws3, R_NET_REV, "NET REVENUE", bold=True)
    write_section(ws3, 37, "REVENUE MIX (% of Gross)")
    write_label(ws3, R_SUB_PCT, "  Subscription %")
    write_label(ws3, R_PPU_PCT, "  Pay-Per-Use %")
    write_label(ws3, R_DFS_PCT, "  DFS %")
    write_label(ws3, R_AD_PCT, "  Advertising %")
    write_label(ws3, R_MRR, "MRR (Subscription Only)", bold=True)
    write_label(ws3, R_ARR, "ARR (= MRR x 12)", bold=True)
    write_label(ws3, R_CUM_REV, "Cumulative Net Revenue", bold=True)

    for m in range(24):
        col = m + 2
        cl = c(col)
        pcl = c(col - 1) if m > 0 else None

        # Subscription revenue — monthly billing is $0 in off-season (users pause/cancel)
        season_check = f'{ug(UG_SEASON, col)}="IN-SEASON"'
        ws3.cell(row=R_PRO_MO, column=col,
                 value=f"=IF({season_check},{ug(UG_END_PRO, col)}*{A['monthly_pct']}*{A['pro_price']},0)")
        ws3.cell(row=R_PRO_AN, column=col,
                 value=f"={ug(UG_END_PRO, col)}*{A['annual_pct']}*{A['pro_price']}")
        ws3.cell(row=R_COMM_MO, column=col,
                 value=f"=IF({season_check},{ug(UG_END_COMM, col)}*{A['monthly_pct']}*{A['comm_price']},0)")
        ws3.cell(row=R_COMM_AN, column=col,
                 value=f"={ug(UG_END_COMM, col)}*{A['annual_pct']}*{A['comm_price']}")
        ws3.cell(row=R_GROSS_SUB, column=col,
                 value=f"={cl}{R_PRO_MO}+{cl}{R_PRO_AN}+{cl}{R_COMM_MO}+{cl}{R_COMM_AN}")
        ws3.cell(row=R_DISC, column=col,
                 value=f"=-({cl}{R_PRO_AN}+{cl}{R_COMM_AN})*{A['annual_disc']}")
        ws3.cell(row=R_NET_SUB, column=col,
                 value=f"={cl}{R_GROSS_SUB}+{cl}{R_DISC}")

        # Pay-per-use (token top-up packs)
        ws3.cell(row=R_AI_FREE, column=col,
                 value=f"={ug(UG_END_FREE, col)}*{A['topup_attach_free']}*{A['topup_price_free']}")
        ws3.cell(row=R_AI_PRO, column=col,
                 value=f"=({ug(UG_END_PRO, col)}+{ug(UG_END_COMM, col)})*{A['topup_attach_paid']}*{A['topup_price_paid']}")
        # Draft kits: months 5,6,17,18 (0-indexed) = columns 7,8,19,20
        ws3.cell(row=R_DRAFT_KIT, column=col,
                 value=f"=IF(OR(COLUMN()=7,COLUMN()=8,COLUMN()=19,COLUMN()=20),{ug(UG_END_TOTAL, col)}*{A['dk_attach']}*{A['dk_price']},0)")
        ws3.cell(row=R_NET_PPU, column=col,
                 value=f"={cl}{R_AI_FREE}+{cl}{R_AI_PRO}+{cl}{R_DRAFT_KIT}")

        # DFS — only in-season AND after DFS launch month
        ws3.cell(row=R_DFS_GROSS, column=col,
                 value=f"=IF(AND(COLUMN()-2>={A['dfs_launch']},{season_check}),{ug(UG_END_TOTAL, col)}*{A['dfs_part']}*{A['dfs_entries']}*{A['dfs_entry']},0)")
        ws3.cell(row=R_DFS_RAKE, column=col,
                 value=f"={cl}{R_DFS_GROSS}*{A['dfs_rake']}")
        ws3.cell(row=R_NET_DFS, column=col,
                 value=f"={cl}{R_DFS_RAKE}")

        # Advertising
        ws3.cell(row=R_AD_IMPS, column=col,
                 value=f"={ug(UG_END_FREE, col)}*IF({ug(UG_SEASON, col)}=\"IN-SEASON\",{A['pv_in']},{A['pv_off']})*{A['ad_elig']}/1000")
        ws3.cell(row=R_AD_REV, column=col,
                 value=f"={cl}{R_AD_IMPS}*{A['ecpm']}")
        ws3.cell(row=R_NET_AD, column=col,
                 value=f"={cl}{R_AD_REV}")

        # Gross revenue
        ws3.cell(row=R_GROSS_REV, column=col,
                 value=f"={cl}{R_NET_SUB}+{cl}{R_NET_PPU}+{cl}{R_NET_DFS}+{cl}{R_NET_AD}")

        # Platform fees
        ws3.cell(row=R_APPLE_FEE, column=col,
                 value=f"={cl}{R_NET_SUB}*{A['mobile_pct']}*{A['ios_split']}*{A['apple_comm']}")
        ws3.cell(row=R_GOOGLE_FEE, column=col,
                 value=f"={cl}{R_NET_SUB}*{A['mobile_pct']}*{A['android_split']}*{A['google_comm']}")
        # Stripe: web subs + PPU + DFS rake
        stripe_elig = f"({cl}{R_NET_SUB}*(1-{A['mobile_pct']})+{cl}{R_NET_PPU}+{cl}{R_NET_DFS})"
        n_txns = f"MAX(1,({ug(UG_END_PRO, col)}+{ug(UG_END_COMM, col)})*0.3+{ug(UG_END_FREE, col)}*{A['topup_attach_free']}+({ug(UG_END_PRO, col)}+{ug(UG_END_COMM, col)})*{A['topup_attach_paid']})"
        ws3.cell(row=R_STRIPE_FEE, column=col,
                 value=f"=IF({stripe_elig}>0,{stripe_elig}*{A['stripe_pct']}+{n_txns}*{A['stripe_flat']},0)")
        ws3.cell(row=R_TOTAL_PF, column=col,
                 value=f"={cl}{R_APPLE_FEE}+{cl}{R_GOOGLE_FEE}+{cl}{R_STRIPE_FEE}")
        ws3.cell(row=R_NET_REV, column=col,
                 value=f"={cl}{R_GROSS_REV}-{cl}{R_TOTAL_PF}")

        # Revenue mix
        ws3.cell(row=R_SUB_PCT, column=col,
                 value=f"=IF({cl}{R_GROSS_REV}>0,{cl}{R_NET_SUB}/{cl}{R_GROSS_REV},0)")
        ws3.cell(row=R_PPU_PCT, column=col,
                 value=f"=IF({cl}{R_GROSS_REV}>0,{cl}{R_NET_PPU}/{cl}{R_GROSS_REV},0)")
        ws3.cell(row=R_DFS_PCT, column=col,
                 value=f"=IF({cl}{R_GROSS_REV}>0,{cl}{R_NET_DFS}/{cl}{R_GROSS_REV},0)")
        ws3.cell(row=R_AD_PCT, column=col,
                 value=f"=IF({cl}{R_GROSS_REV}>0,{cl}{R_NET_AD}/{cl}{R_GROSS_REV},0)")

        # MRR / ARR
        ws3.cell(row=R_MRR, column=col, value=f"={cl}{R_NET_SUB}")
        ws3.cell(row=R_ARR, column=col, value=f"={cl}{R_NET_SUB}*12")

        # Cumulative net revenue
        if m == 0:
            ws3.cell(row=R_CUM_REV, column=col, value=f"={cl}{R_NET_REV}")
        else:
            ws3.cell(row=R_CUM_REV, column=col, value=f"={pcl}{R_CUM_REV}+{cl}{R_NET_REV}")

    # Apply formats to Revenue
    for r in [R_PRO_MO, R_PRO_AN, R_COMM_MO, R_COMM_AN, R_GROSS_SUB, R_DISC,
              R_AI_FREE, R_AI_PRO, R_DRAFT_KIT, R_DFS_GROSS, R_DFS_RAKE, R_AD_REV,
              R_APPLE_FEE, R_GOOGLE_FEE, R_STRIPE_FEE, R_TOTAL_PF]:
        style_formula_row(ws3, r, 2, 25, usd_fmt)
    for r in [R_NET_SUB, R_NET_PPU, R_NET_DFS, R_NET_AD, R_GROSS_REV, R_NET_REV, R_MRR, R_ARR, R_CUM_REV]:
        style_formula_row(ws3, r, 2, 25, usd_fmt, is_total=True)
    for r in [R_SUB_PCT, R_PPU_PCT, R_DFS_PCT, R_AD_PCT]:
        style_formula_row(ws3, r, 2, 25, pct_fmt)
    style_formula_row(ws3, R_AD_IMPS, 2, 25, num_fmt)

    # Conditional formatting for negative values on Revenue
    ws3.conditional_formatting.add(
        f"B{R_DISC}:Y{R_DISC}",
        CellIsRule(operator="lessThan", formula=["0"], font=neg_font))


    # ──────────────────────────────────────────────────────────────────
    # TAB 4: COGS & OPERATING EXPENSES — All formulas
    # ──────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("COGS & Operating Expenses")
    set_col_widths(ws4, [38] + [14] * 24)
    write_header_row(ws4, 1, ["Cost Line"] + MONTHS)

    # Labels
    write_section(ws4, 2, "COST OF GOODS SOLD (COGS)")
    write_label(ws4, CE_SUPABASE, "  Supabase (Database + Auth)")
    write_label(ws4, CE_FIREBASE, "  Firebase Hosting")
    write_label(ws4, CE_CLAUDE, "  Claude API (Stormy AI)")
    write_label(ws4, CE_PROXY, "  Proxy Rotation (Data Pipeline)")
    write_label(ws4, CE_DOMAIN, "  Domain / SSL / Misc")
    write_label(ws4, CE_TOTAL_COGS, "  TOTAL COGS", bold=True)
    write_label(ws4, CE_GROSS_PROFIT, "GROSS PROFIT", bold=True)
    write_label(ws4, CE_GROSS_MARGIN, "GROSS MARGIN %")
    write_section(ws4, 13, "OPERATING EXPENSES (OPEX)")
    write_section(ws4, 14, "  SALES & MARKETING")
    write_label(ws4, CE_DIG_MKT, "    Digital Marketing")
    write_label(ws4, CE_CONT_MKT, "    Influencer / Content Marketing")
    write_label(ws4, CE_TOTAL_SM, "    Total Sales & Marketing", bold=True)
    write_section(ws4, 19, "  GENERAL & ADMINISTRATIVE")
    write_label(ws4, CE_TEAM, "    Contractors / Team")
    write_label(ws4, CE_LEGAL, "    Legal & Accounting (Basic)")
    write_label(ws4, CE_DFS_REG, "    DFS Regulatory — Canada")
    write_label(ws4, CE_US_DFS_LIC, "    US DFS Licensing (State-by-State)")
    write_label(ws4, CE_US_DFS_COMPLY, "    US DFS Compliance Overhead")
    write_label(ws4, CE_US_DFS_TOTAL, "    Total US DFS Expansion", bold=True)
    write_label(ws4, CE_APPLE_DEV, "    Apple Developer Program")
    write_label(ws4, CE_GOOGLE_REG, "    Google Play Registration")
    write_label(ws4, CE_TOTAL_GA, "    Total G&A", bold=True)
    write_section(ws4, 30, "  CONFERENCE & TRAVEL")
    write_label(ws4, CE_CONF, "    Conference Costs", bold=True)
    write_label(ws4, CE_TOTAL_CONF, "    Total Conference & Travel", bold=True)
    write_label(ws4, CE_TOTAL_OPEX, "  TOTAL OPEX", bold=True)
    write_label(ws4, CE_EBITDA, "EBITDA", bold=True)
    write_label(ws4, CE_EBITDA_MARGIN, "EBITDA MARGIN %")
    write_label(ws4, CE_NET_INCOME, "NET INCOME", bold=True)

    for m in range(24):
        col = m + 2
        cl = c(col)

        # COGS
        ws4.cell(row=CE_SUPABASE, column=col,
                 value=f"=IF({ug(UG_END_TOTAL, col)}>{A['sb_thresh']},{A['sb_team']},{A['sb_pro']})")
        ws4.cell(row=CE_FIREBASE, column=col,
                 value=f"={A['fb_base']}+MAX(0,{ug(UG_END_TOTAL, col)}-1000)*{A['fb_scale']}")
        # Claude API cost = (users × tokens/mo per tier) × cost per 1M tokens / 1,000,000
        ws4.cell(row=CE_CLAUDE, column=col,
                 value=f"=({ug(UG_END_FREE, col)}*{A['tok_free']}+{ug(UG_END_PRO, col)}*{A['tok_pro']}+{ug(UG_END_COMM, col)}*{A['tok_comm']})/1000000*{A['cl_cost_per_m']}")
        ws4.cell(row=CE_PROXY, column=col, value=f"={A['proxy']}")
        ws4.cell(row=CE_DOMAIN, column=col, value=f"={A['domain']}")
        ws4.cell(row=CE_TOTAL_COGS, column=col,
                 value=f"={cl}{CE_SUPABASE}+{cl}{CE_FIREBASE}+{cl}{CE_CLAUDE}+{cl}{CE_PROXY}+{cl}{CE_DOMAIN}")

        # Gross profit
        ws4.cell(row=CE_GROSS_PROFIT, column=col,
                 value=f"={rev(R_NET_REV, col)}-{cl}{CE_TOTAL_COGS}")
        ws4.cell(row=CE_GROSS_MARGIN, column=col,
                 value=f"=IF({rev(R_NET_REV, col)}<>0,{cl}{CE_GROSS_PROFIT}/{rev(R_NET_REV, col)},0)")

        # OPEX - Sales & Marketing (gated on marketing launch month)
        ws4.cell(row=CE_DIG_MKT, column=col,
                 value=f"=IF(COLUMN()-2>={A['mkt_launch']},{ug(UG_NEW, col)}*{A['cac']},0)")
        ws4.cell(row=CE_CONT_MKT, column=col,
                 value=f"={A[f'cont_mkt_{m}']}")
        ws4.cell(row=CE_TOTAL_SM, column=col,
                 value=f"={cl}{CE_DIG_MKT}+{cl}{CE_CONT_MKT}")

        # G&A
        # Team cost: step function based on month index (COLUMN()-2)
        ws4.cell(row=CE_TEAM, column=col,
                 value=f"=IF(COLUMN()-2<6,{A['team_1_6']},IF(COLUMN()-2<12,{A['team_7_12']},IF(COLUMN()-2<18,{A['team_13_18']},{A['team_19_24']})))")
        ws4.cell(row=CE_LEGAL, column=col,
                 value=f"=IF(COLUMN()<14,{A['legal_y1']},{A['legal_y2']})")
        # DFS Regulatory (Canada): $0 until total users exceed threshold
        ws4.cell(row=CE_DFS_REG, column=col,
                 value=f"=IF({ug(UG_END_TOTAL, col)}>={A['dfs_reg_thresh']},{A['dfs_legal_opinion']}+{A['dfs_compliance']}+{A['dfs_insurance']}+{A['dfs_player_trust']},0)")

        # US DFS Expansion: state-by-state licensing (toggle on us_expand_month)
        # Month 0 of expansion = Year 1 fees; subsequent months = annual/12
        expand_active = f"COLUMN()-2>={A['us_expand_month']}"
        months_since = f"COLUMN()-2-{A['us_expand_month']}"
        # Build state licensing sum: IF first 12 months → Y1/12, ELSE → annual/12
        y1_monthly_parts = []
        ann_monthly_parts = []
        for st in US_DFS_STATES:
            y1_monthly_parts.append(f"{A[f'us_{st}_y1']}/12")
            ann_monthly_parts.append(f"{A[f'us_{st}_ann']}/12")
        y1_sum = "+".join(y1_monthly_parts)
        ann_sum = "+".join(ann_monthly_parts)
        ws4.cell(row=CE_US_DFS_LIC, column=col,
                 value=f"=IF({expand_active},IF({months_since}<12,{y1_sum},{ann_sum}),0)")

        # US compliance overhead (annual costs / 12, only when expansion active)
        comply_parts = "+".join([
            f"{A['us_legal_counsel']}/12",
            f"{A['us_geolocation']}/12",
            f"{A['us_audit']}/12",
            f"{A['us_responsible_gaming']}/12",
            f"{A['us_reporting']}/12",
            f"{A['us_bg_checks']}/12",
        ])
        # NY revenue tax: DFS rake revenue * NY share * NY tax rate
        ny_rev_tax = f"{rev(R_NET_DFS, col)}*{A['us_ny_rev_share']}*{A['us_ny_rev_tax']}"
        ws4.cell(row=CE_US_DFS_COMPLY, column=col,
                 value=f"=IF({expand_active},{comply_parts}+{ny_rev_tax},0)")

        # Total US DFS
        ws4.cell(row=CE_US_DFS_TOTAL, column=col,
                 value=f"={cl}{CE_US_DFS_LIC}+{cl}{CE_US_DFS_COMPLY}")

        ws4.cell(row=CE_APPLE_DEV, column=col,
                 value=f"={A['apple_dev']}/12")
        ws4.cell(row=CE_GOOGLE_REG, column=col,
                 value=f"=IF(COLUMN()=2,{A['google_reg']},0)")
        ws4.cell(row=CE_TOTAL_GA, column=col,
                 value=f"={cl}{CE_TEAM}+{cl}{CE_LEGAL}+{cl}{CE_DFS_REG}+{cl}{CE_US_DFS_TOTAL}+{cl}{CE_APPLE_DEV}+{cl}{CE_GOOGLE_REG}")

        # Conference costs (INPUT — yellow editable cells)
        conf_val = CONFERENCE_COSTS.get(m, 0)
        cell = ws4.cell(row=CE_CONF, column=col, value=conf_val)
        cell.fill = input_fill
        cell.font = input_font

        ws4.cell(row=CE_TOTAL_CONF, column=col, value=f"={cl}{CE_CONF}")

        # Total OPEX
        ws4.cell(row=CE_TOTAL_OPEX, column=col,
                 value=f"={cl}{CE_TOTAL_SM}+{cl}{CE_TOTAL_GA}+{cl}{CE_TOTAL_CONF}")

        # EBITDA
        ws4.cell(row=CE_EBITDA, column=col,
                 value=f"={cl}{CE_GROSS_PROFIT}-{cl}{CE_TOTAL_OPEX}")
        ws4.cell(row=CE_EBITDA_MARGIN, column=col,
                 value=f"=IF({rev(R_NET_REV, col)}<>0,{cl}{CE_EBITDA}/{rev(R_NET_REV, col)},0)")
        ws4.cell(row=CE_NET_INCOME, column=col, value=f"={cl}{CE_EBITDA}")

    # Apply formats to COGS
    for r in [CE_SUPABASE, CE_PROXY, CE_DOMAIN, CE_DIG_MKT, CE_CONT_MKT,
              CE_TEAM, CE_LEGAL, CE_DFS_REG, CE_US_DFS_LIC, CE_US_DFS_COMPLY, CE_GOOGLE_REG, CE_CONF]:
        style_formula_row(ws4, r, 2, 25, usd_fmt)
    for r in [CE_FIREBASE, CE_CLAUDE, CE_APPLE_DEV]:
        style_formula_row(ws4, r, 2, 25, usd_cents_fmt)
    for r in [CE_TOTAL_COGS, CE_TOTAL_SM, CE_US_DFS_TOTAL, CE_TOTAL_GA, CE_TOTAL_CONF, CE_TOTAL_OPEX]:
        style_formula_row(ws4, r, 2, 25, usd_fmt, is_total=True)
    for r in [CE_GROSS_PROFIT, CE_EBITDA, CE_NET_INCOME]:
        style_formula_row(ws4, r, 2, 25, usd_fmt, is_total=True)
    for r in [CE_GROSS_MARGIN, CE_EBITDA_MARGIN]:
        style_formula_row(ws4, r, 2, 25, pct_fmt)

    # Conditional formatting: green for positive EBITDA/profit, red for negative
    for r in [CE_GROSS_PROFIT, CE_EBITDA, CE_NET_INCOME]:
        ws4.conditional_formatting.add(
            f"B{r}:Y{r}",
            CellIsRule(operator="greaterThan", formula=["0"], font=pos_font))
        ws4.conditional_formatting.add(
            f"B{r}:Y{r}",
            CellIsRule(operator="lessThan", formula=["0"], font=neg_font))


    # ──────────────────────────────────────────────────────────────────
    # TAB 5: UNIT ECONOMICS — All formulas
    # ──────────────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Unit Economics")
    set_col_widths(ws5, [38] + [14] * 24)
    write_header_row(ws5, 1, ["Metric"] + MONTHS)

    write_label(ws5, UE_ARPU_TOTAL, "ARPU (Total) — Net Rev / Users")
    write_label(ws5, UE_ARPU_PAID, "ARPU (Paid) — Sub Rev / Paid")
    write_label(ws5, UE_PAID_SUBS, "Paid Subscribers")
    write_label(ws5, UE_BLEND_CHURN, "Blended Monthly Churn (Paid)")
    write_label(ws5, UE_LTV, "Implied LTV (ARPU / Churn)")
    write_label(ws5, UE_CAC, "CAC (S&M / New Users)")
    write_label(ws5, UE_LTV_CAC, "LTV:CAC Ratio")
    write_label(ws5, UE_PAYBACK, "Payback Period (months)")
    write_label(ws5, UE_TEAM_SIZE, "Team Size (approx)")
    write_label(ws5, UE_REV_EMP, "Revenue Per Employee")

    for m in range(24):
        col = m + 2
        cl = c(col)

        ws5.cell(row=UE_ARPU_TOTAL, column=col,
                 value=f"=IF({ug(UG_END_TOTAL, col)}>0,{rev(R_NET_REV, col)}/{ug(UG_END_TOTAL, col)},0)")
        ws5.cell(row=UE_PAID_SUBS, column=col,
                 value=f"={ug(UG_END_PRO, col)}+{ug(UG_END_COMM, col)}")
        ws5.cell(row=UE_ARPU_PAID, column=col,
                 value=f"=IF({cl}{UE_PAID_SUBS}>0,{rev(R_NET_SUB, col)}/{cl}{UE_PAID_SUBS},0)")
        # Blended paid churn
        paid_churned = f"{ug(UG_CHURNED_PRO, col)}+{ug(UG_CHURNED_COMM, col)}"
        paid_base = f"{cl}{UE_PAID_SUBS}+{ug(UG_CHURNED_PRO, col)}+{ug(UG_CHURNED_COMM, col)}"
        ws5.cell(row=UE_BLEND_CHURN, column=col,
                 value=f"=IF({paid_base}>0,({paid_churned})/({paid_base}),0)")
        ws5.cell(row=UE_LTV, column=col,
                 value=f"=IF({cl}{UE_BLEND_CHURN}>0,{cl}{UE_ARPU_PAID}/{cl}{UE_BLEND_CHURN},0)")
        ws5.cell(row=UE_CAC, column=col,
                 value=f"=IF({ug(UG_NEW, col)}>0,{cogs(CE_TOTAL_SM, col)}/{ug(UG_NEW, col)},0)")
        ws5.cell(row=UE_LTV_CAC, column=col,
                 value=f"=IF({cl}{UE_CAC}>0,{cl}{UE_LTV}/{cl}{UE_CAC},0)")
        ws5.cell(row=UE_PAYBACK, column=col,
                 value=f"=IF({cl}{UE_ARPU_PAID}>0,{cl}{UE_CAC}/{cl}{UE_ARPU_PAID},0)")
        ws5.cell(row=UE_TEAM_SIZE, column=col,
                 value=f"=IF(COLUMN()-2<6,0,IF(COLUMN()-2<12,1,IF(COLUMN()-2<18,2,3)))")
        ws5.cell(row=UE_REV_EMP, column=col,
                 value=f"=IF({cl}{UE_TEAM_SIZE}>0,{rev(R_NET_REV, col)}/{cl}{UE_TEAM_SIZE},0)")

    # Formats
    style_formula_row(ws5, UE_ARPU_TOTAL, 2, 25, usd_cents_fmt)
    style_formula_row(ws5, UE_ARPU_PAID, 2, 25, usd_cents_fmt)
    style_formula_row(ws5, UE_PAID_SUBS, 2, 25, num_fmt)
    style_formula_row(ws5, UE_BLEND_CHURN, 2, 25, pct_fmt)
    style_formula_row(ws5, UE_LTV, 2, 25, usd_fmt)
    style_formula_row(ws5, UE_CAC, 2, 25, usd_cents_fmt)
    style_formula_row(ws5, UE_LTV_CAC, 2, 25, '0.0"x"')
    style_formula_row(ws5, UE_PAYBACK, 2, 25, "0.0")
    style_formula_row(ws5, UE_TEAM_SIZE, 2, 25, num_fmt)
    style_formula_row(ws5, UE_REV_EMP, 2, 25, usd_fmt)

    # ──────────────────────────────────────────────────────────────────
    # TAB 6: CASH FLOW & RUNWAY — All formulas
    # ──────────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("Cash Flow & Runway")
    set_col_widths(ws6, [38] + [14] * 24)
    write_header_row(ws6, 1, ["Cash Flow"] + MONTHS)

    write_label(ws6, CF_START, "Starting Cash", bold=True)
    write_label(ws6, CF_MONTHLY, "Monthly Net Income (EBITDA)")
    write_label(ws6, CF_CUM, "Cumulative Cash Position", bold=True)
    write_label(ws6, CF_BURN, "Monthly Burn Rate")
    write_label(ws6, CF_RUNWAY, "Runway (months)")

    for m in range(24):
        col = m + 2
        cl = c(col)
        pcl = c(col - 1) if m > 0 else None

        # Starting cash only in month 0
        if m == 0:
            ws6.cell(row=CF_START, column=col, value=f"={A['start_cash']}")
        # else leave blank

        ws6.cell(row=CF_MONTHLY, column=col, value=f"={cogs(CE_EBITDA, col)}")

        if m == 0:
            ws6.cell(row=CF_CUM, column=col,
                     value=f"={A['start_cash']}+{cl}{CF_MONTHLY}")
        else:
            ws6.cell(row=CF_CUM, column=col,
                     value=f"={pcl}{CF_CUM}+{cl}{CF_MONTHLY}")

        ws6.cell(row=CF_BURN, column=col,
                 value=f"=IF({cl}{CF_MONTHLY}<0,-{cl}{CF_MONTHLY},0)")
        ws6.cell(row=CF_RUNWAY, column=col,
                 value=f"=IF(AND({cl}{CF_BURN}>0,{cl}{CF_CUM}>0),{cl}{CF_CUM}/{cl}{CF_BURN},0)")

    # Formats
    style_formula_row(ws6, CF_START, 2, 25, usd_fmt)
    style_formula_row(ws6, CF_MONTHLY, 2, 25, usd_fmt)
    style_formula_row(ws6, CF_CUM, 2, 25, usd_fmt, is_total=True)
    style_formula_row(ws6, CF_BURN, 2, 25, usd_fmt)
    style_formula_row(ws6, CF_RUNWAY, 2, 25, "0.0")

    # Conditional formatting
    ws6.conditional_formatting.add(
        f"B{CF_CUM}:Y{CF_CUM}",
        CellIsRule(operator="greaterThan", formula=["0"], font=pos_font))
    ws6.conditional_formatting.add(
        f"B{CF_CUM}:Y{CF_CUM}",
        CellIsRule(operator="lessThan", formula=["0"], font=neg_font))
    ws6.conditional_formatting.add(
        f"B{CF_MONTHLY}:Y{CF_MONTHLY}",
        CellIsRule(operator="lessThan", formula=["0"], font=neg_font))

    # Cash flow positive month indicator
    r = CF_RUNWAY + 2
    ws6.cell(row=r, column=1, value="Cash Flow Positive Month:").font = subsection_font
    # This uses a formula to find the first profitable month
    ws6.cell(row=r, column=2,
             value=f'=IFERROR(INDEX(B1:Y1,MATCH(TRUE,B{CF_MONTHLY}:Y{CF_MONTHLY}>0,0)),"Not within 24 months")')
    ws6.cell(row=r, column=2).font = Font(bold=True, color=GREEN)


    # ──────────────────────────────────────────────────────────────────
    # TAB 7: ANNUAL SUMMARY — SUM formulas over Y1/Y2
    # ──────────────────────────────────────────────────────────────────
    ws7 = wb.create_sheet("Annual Summary")
    set_col_widths(ws7, [38, 18, 18, 18])
    write_header_row(ws7, 1, ["Metric", "Year 1 (Apr 26-Mar 27)", "Year 2 (Apr 27-Mar 28)", "YoY Growth %"])

    # Helper: Y1=B:M, Y2=N:Y on various sheets
    def annual_row(ws_out, out_row, label, sheet_ref_func, src_row, is_sum=True, is_section=False, fmt=usd_fmt):
        ws_out.cell(row=out_row, column=1, value=label)
        if is_section:
            ws_out.cell(row=out_row, column=1).font = section_font
            return
        ws_out.cell(row=out_row, column=1).font = Font(size=10)
        if is_sum:
            # SUM across months
            ws_out.cell(row=out_row, column=2,
                        value=f"=SUM({sheet_ref_func(src_row, 2)}:{sheet_ref_func(src_row, 13)})")
            ws_out.cell(row=out_row, column=3,
                        value=f"=SUM({sheet_ref_func(src_row, 14)}:{sheet_ref_func(src_row, 25)})")
        else:
            # Point-in-time (end of year)
            ws_out.cell(row=out_row, column=2, value=f"={sheet_ref_func(src_row, 13)}")  # M = month 11
            ws_out.cell(row=out_row, column=3, value=f"={sheet_ref_func(src_row, 25)}")  # Y = month 23
        # YoY growth
        ws_out.cell(row=out_row, column=4,
                    value=f"=IF(ABS(B{out_row})>0,(C{out_row}-B{out_row})/ABS(B{out_row}),0)")
        for col in [2, 3]:
            ws_out.cell(row=out_row, column=col).number_format = fmt
            ws_out.cell(row=out_row, column=col).alignment = Alignment(horizontal="right")
        ws_out.cell(row=out_row, column=4).number_format = pct_fmt
        ws_out.cell(row=out_row, column=4).alignment = Alignment(horizontal="right")

    r = 2
    annual_row(ws7, r, "KEY METRICS", None, None, is_section=True); r += 1
    annual_row(ws7, r, "End-of-Year Total Users", ug, UG_END_TOTAL, is_sum=False, fmt=num_fmt); r += 1
    annual_row(ws7, r, "End-of-Year Pro Users", ug, UG_END_PRO, is_sum=False, fmt=num_fmt); r += 1
    annual_row(ws7, r, "End-of-Year Commissioner Users", ug, UG_END_COMM, is_sum=False, fmt=num_fmt); r += 1
    r += 1  # blank
    annual_row(ws7, r, "REVENUE", None, None, is_section=True); r += 1
    annual_row(ws7, r, "Net Subscription Revenue", rev, R_NET_SUB); r += 1
    annual_row(ws7, r, "Pay-Per-Use Revenue", rev, R_NET_PPU); r += 1
    annual_row(ws7, r, "DFS Rake Revenue", rev, R_NET_DFS); r += 1
    annual_row(ws7, r, "Advertising Revenue", rev, R_NET_AD); r += 1
    annual_row(ws7, r, "Gross Revenue", rev, R_GROSS_REV); r += 1
    annual_row(ws7, r, "Platform Fees", rev, R_TOTAL_PF); r += 1
    annual_row(ws7, r, "Net Revenue", rev, R_NET_REV); r += 1
    net_rev_y1_row = r - 1  # remember for margin calc
    r += 1  # blank
    annual_row(ws7, r, "COSTS", None, None, is_section=True); r += 1
    annual_row(ws7, r, "Total COGS", cogs, CE_TOTAL_COGS); r += 1
    annual_row(ws7, r, "Gross Profit", cogs, CE_GROSS_PROFIT); r += 1
    gp_row = r - 1
    annual_row(ws7, r, "Sales & Marketing", cogs, CE_TOTAL_SM); r += 1
    annual_row(ws7, r, "DFS Regulatory — Canada", cogs, CE_DFS_REG); r += 1
    annual_row(ws7, r, "US DFS Expansion (Total)", cogs, CE_US_DFS_TOTAL); r += 1
    annual_row(ws7, r, "Total OPEX", cogs, CE_TOTAL_OPEX); r += 1
    annual_row(ws7, r, "EBITDA / Net Income", cogs, CE_EBITDA); r += 1
    ebitda_row = r - 1
    r += 1  # blank
    annual_row(ws7, r, "MARGINS", None, None, is_section=True); r += 1

    # Gross Margin %
    ws7.cell(row=r, column=1, value="Gross Margin %").font = Font(size=10)
    ws7.cell(row=r, column=2, value=f"=IF(B{net_rev_y1_row}<>0,B{gp_row}/B{net_rev_y1_row},0)")
    ws7.cell(row=r, column=3, value=f"=IF(C{net_rev_y1_row}<>0,C{gp_row}/C{net_rev_y1_row},0)")
    for col in [2, 3]:
        ws7.cell(row=r, column=col).number_format = pct_fmt
        ws7.cell(row=r, column=col).alignment = Alignment(horizontal="right")
    r += 1

    # EBITDA Margin %
    ws7.cell(row=r, column=1, value="EBITDA Margin %").font = Font(size=10)
    ws7.cell(row=r, column=2, value=f"=IF(B{net_rev_y1_row}<>0,B{ebitda_row}/B{net_rev_y1_row},0)")
    ws7.cell(row=r, column=3, value=f"=IF(C{net_rev_y1_row}<>0,C{ebitda_row}/C{net_rev_y1_row},0)")
    for col in [2, 3]:
        ws7.cell(row=r, column=col).number_format = pct_fmt
        ws7.cell(row=r, column=col).alignment = Alignment(horizontal="right")

    # Conditional formatting for EBITDA
    ws7.conditional_formatting.add(
        f"B{ebitda_row}:C{ebitda_row}",
        CellIsRule(operator="greaterThan", formula=["0"], font=pos_font))
    ws7.conditional_formatting.add(
        f"B{ebitda_row}:C{ebitda_row}",
        CellIsRule(operator="lessThan", formula=["0"], font=neg_font))


    # ──────────────────────────────────────────────────────────────────
    # TAB 8: SENSITIVITY ANALYSIS — Formula-based (dynamic)
    # ──────────────────────────────────────────────────────────────────
    ws8 = wb.create_sheet("Sensitivity Analysis")
    set_col_widths(ws8, [35, 16, 16, 16, 16])

    # Common references for Y2 revenue components
    y2_sub = "SUM(Revenue!N9:Y9)"
    y2_ppu = "SUM(Revenue!N15:Y15)"
    y2_dfs = "SUM(Revenue!N20:Y20)"
    y2_ad = "SUM(Revenue!N25:Y25)"
    y2_gross = "SUM(Revenue!N27:Y27)"

    # ── Conversion Rate Sensitivity ──
    r = 2
    ws8.cell(row=r, column=1, value="CONVERSION RATE SENSITIVITY (Y2 Pro)").font = section_font
    r += 1
    write_header_row(ws8, r, ["Conv Rate", "Y2 Gross Rev (est)", "vs Base", "", ""], 1)
    r += 1

    for cv in [0.02, 0.035, 0.055, 0.08, 0.10]:
        ws8.cell(row=r, column=1, value=cv)
        ws8.cell(row=r, column=1).number_format = pct_fmt
        ws8.cell(row=r, column=1).fill = input_fill
        ws8.cell(row=r, column=1).font = input_font
        # Y2 gross = (test/base) * y2_sub + y2_ppu + y2_dfs + y2_ad
        ws8.cell(row=r, column=2,
                 value=f"=(A{r}/{A['conv_pro_y2']})*{y2_sub}+{y2_ppu}+{y2_dfs}+{y2_ad}")
        ws8.cell(row=r, column=2).number_format = usd_fmt
        ws8.cell(row=r, column=3,
                 value=f"=IF({y2_gross}<>0,(B{r}-{y2_gross})/{y2_gross},0)")
        ws8.cell(row=r, column=3).number_format = pct_fmt
        if cv == 0.055:
            ws8.cell(row=r, column=4, value="<-- BASE").font = Font(bold=True, color=GREEN)
        r += 1

    # ── Churn Rate Sensitivity ──
    r += 1
    ws8.cell(row=r, column=1, value="CHURN RATE SENSITIVITY (Paid In-Season)").font = section_font
    r += 1
    write_header_row(ws8, r, ["Monthly Churn", "Y2 Gross Rev (est)", "vs Base", "", ""], 1)
    r += 1

    for ch in [0.02, 0.04, 0.06, 0.09, 0.12]:
        ws8.cell(row=r, column=1, value=ch)
        ws8.cell(row=r, column=1).number_format = pct_fmt
        ws8.cell(row=r, column=1).fill = input_fill
        ws8.cell(row=r, column=1).font = input_font
        # Retention ratio compounded over 6 in-season months
        ws8.cell(row=r, column=2,
                 value=f"=((1-A{r})/(1-{A['ch_pro_in']}))^6*{y2_sub}+{y2_ppu}+{y2_dfs}+{y2_ad}")
        ws8.cell(row=r, column=2).number_format = usd_fmt
        ws8.cell(row=r, column=3,
                 value=f"=IF({y2_gross}<>0,(B{r}-{y2_gross})/{y2_gross},0)")
        ws8.cell(row=r, column=3).number_format = pct_fmt
        if ch == 0.04:
            ws8.cell(row=r, column=4, value="<-- BASE").font = Font(bold=True, color=GREEN)
        r += 1

    # ── User Growth Scenarios ──
    r += 1
    ws8.cell(row=r, column=1, value="USER GROWTH SCENARIOS").font = section_font
    r += 1
    write_header_row(ws8, r, ["Scenario", "Y2 End Users (est)", "24-Mo Gross Rev (est)", "vs Base", ""], 1)
    r += 1

    total_gross = "SUM(Revenue!B27:Y27)"
    y2_end_users = f"'User Growth'!Y{UG_END_TOTAL}"
    for scenario, mult in [("Conservative (0.5x)", 0.5), ("Base (1x)", 1.0), ("Aggressive (2x)", 2.0), ("Hypergrowth (5x)", 5.0)]:
        ws8.cell(row=r, column=1, value=scenario)
        if mult == 1.0:
            ws8.cell(row=r, column=1).font = Font(bold=True, color=GREEN)
        ws8.cell(row=r, column=2, value=f"=ROUND({y2_end_users}*{mult},0)")
        ws8.cell(row=r, column=2).number_format = num_fmt
        ws8.cell(row=r, column=3, value=f"=ROUND({total_gross}*{mult},0)")
        ws8.cell(row=r, column=3).number_format = usd_fmt
        ws8.cell(row=r, column=4, value=f"=IF({total_gross}<>0,(C{r}-{total_gross})/{total_gross},0)")
        ws8.cell(row=r, column=4).number_format = pct_fmt
        if mult == 1.0:
            ws8.cell(row=r, column=5, value="<-- BASE").font = Font(bold=True, color=GREEN)
        r += 1

    # ── Pricing Sensitivity ──
    r += 1
    ws8.cell(row=r, column=1, value="PRICING SENSITIVITY (Pro Tier)").font = section_font
    r += 1
    write_header_row(ws8, r, ["Pro Price", "Y2 Sub Rev (est)", "vs Base", "", ""], 1)
    r += 1

    for price in [3.99, 4.99, 6.99, 9.99]:
        ws8.cell(row=r, column=1, value=price)
        ws8.cell(row=r, column=1).number_format = "$#,##0.00"
        ws8.cell(row=r, column=1).fill = input_fill
        ws8.cell(row=r, column=1).font = input_font
        # Price ratio * conversion adjustment * Y2 sub rev
        ws8.cell(row=r, column=2,
                 value=f"=ROUND((A{r}/{A['pro_price']})*IF(A{r}<={A['pro_price']},1,MAX(0.7,1-(A{r}-{A['pro_price']})*0.04))*{y2_sub},0)")
        ws8.cell(row=r, column=2).number_format = usd_fmt
        ws8.cell(row=r, column=3,
                 value=f"=IF({y2_sub}<>0,(B{r}-{y2_sub})/{y2_sub},0)")
        ws8.cell(row=r, column=3).number_format = pct_fmt
        if price == 4.99:
            ws8.cell(row=r, column=4, value="<-- BASE").font = Font(bold=True, color=GREEN)
        r += 1

    # Conditional formatting for sensitivity vs Base columns
    ws8.conditional_formatting.add(
        "C4:C30",
        CellIsRule(operator="lessThan", formula=["0"], font=neg_font))
    ws8.conditional_formatting.add(
        "C4:C30",
        CellIsRule(operator="greaterThan", formula=["0"], font=pos_font))
    ws8.conditional_formatting.add(
        "D4:D30",
        CellIsRule(operator="lessThan", formula=["0"], font=neg_font))
    ws8.conditional_formatting.add(
        "D4:D30",
        CellIsRule(operator="greaterThan", formula=["0"], font=pos_font))


    # ──────────────────────────────────────────────────────────────────
    # TAB 9: SOURCES & METHODOLOGY
    # ──────────────────────────────────────────────────────────────────
    ws9 = wb.create_sheet("Sources & Methodology")
    set_col_widths(ws9, [45, 45, 75])
    write_header_row(ws9, 1, ["Assumption / Data Point", "Source Name", "URL"])

    sources = [
        ("CONVERSION RATES", "", ""),
        ("Freemium conversion benchmarks", "OpenView / Lenny's Newsletter", "https://www.lennysnewsletter.com/p/what-is-a-good-free-to-paid-conversion"),
        ("Mobile freemium median 2.18%", "RevenueCat State of Subscription Apps 2025", "https://www.revenuecat.com/state-of-subscription-apps-2025/"),
        ("", "", ""),
        ("CHURN RATES", "", ""),
        ("Subscription churn benchmarks", "MarketingLTB Subscription Statistics", "https://marketingltb.com/blog/statistics/subscription-statistics/"),
        ("Sleeper 75% annual retention", "Sleeper Series C PR", "https://www.prnewswire.com/news-releases/fantasy-sports-app-sleeper-raises-40-million-doubles-user-base-in-2021-301379179.html"),
        ("", "", ""),
        ("DFS RAKE", "", ""),
        ("DraftKings/FanDuel rake 10-16%", "Legal Sports Report", "https://www.legalsportsreport.com/15721/draftkings-fanduel-rake-increases/"),
        ("Rake comparison data", "RotoPicks", "http://www.rotopicks.com/fantasy-leagues/rake-comparison.php"),
        ("", "", ""),
        ("ADVERTISING", "", ""),
        ("Mobile ad CPM rates", "BusinessOfApps", "https://www.businessofapps.com/ads/research/mobile-app-advertising-cpm-rates/"),
        ("eCPM benchmarks by format", "Appodeal eCPM Report", "https://appodeal.com/blog/mobile-ecpm-report-app-ad-monetization-worldwide-performance/"),
        ("", "", ""),
        ("INFRASTRUCTURE", "", ""),
        ("Supabase pricing tiers", "Supabase Pricing", "https://supabase.com/pricing"),
        ("Supabase true cost analysis", "Metacto Blog", "https://www.metacto.com/blogs/the-true-cost-of-supabase-a-comprehensive-guide-to-pricing-integration-and-maintenance"),
        ("Firebase Hosting pricing", "Firebase Docs", "https://firebase.google.com/docs/hosting/usage-quotas-pricing"),
        ("Claude API pricing (Sonnet)", "Anthropic Pricing — $3/M in, $15/M out, $0.30/M cached", "https://docs.anthropic.com/en/docs/about-claude/models"),
        ("", "", ""),
        ("PLATFORM FEES", "", ""),
        ("Apple Small Business Program (15%)", "Apple Developer", "https://developer.apple.com/app-store/small-business-program/"),
        ("Apple SBP implementation", "RevenueCat Engineering Blog", "https://www.revenuecat.com/blog/engineering/small-business-program/"),
        ("Google Play reduced fee (15%)", "Google Play Console", "https://play.google.com/console/about/programs/reducedservicefee/"),
        ("Stripe processing fees", "Stripe Pricing", "https://stripe.com/pricing"),
        ("Apple Developer Program $99/yr", "Apple Developer", "https://developer.apple.com/support/enrollment/"),
        ("", "", ""),
        ("DFS COMPLIANCE — CANADA (NO LICENSE REQUIRED)", "", ""),
        ("Criminal Code s.206(1)(d) skill exemption", "DFS is LEGAL — contest of skill, no gaming license", "https://laws-lois.justice.gc.ca/eng/acts/c-46/page-46.html"),
        ("NOT regulated by iGaming Ontario", "iGO = sportsbooks/casinos only, not DFS", "https://igamingontario.ca/en/operators"),
        ("NOT regulated by AGCO for DFS", "AGCO gaming regs do not cover skill contests", "https://www.agco.ca/igaming"),
        ("DK/FD DFS precedent in Canada", "Both operate DFS in Canada without gaming license", ""),
        ("FINTRAC AML if prize pools >$10K", "Reporting obligations — not a license, just compliance", "https://fintrac-canafe.canada.ca/re-ed/intro-eng"),
        ("Legal opinion $3K-8K one-time", "Confirms s.206(1)(d) applies — amortized in model", ""),
        ("", "", ""),
        ("MARKETING / CAC", "", ""),
        ("DraftKings/FanDuel CAC $200-350", "BusinessOfApps UA Costs", "https://www.businessofapps.com/marketplace/user-acquisition/research/user-acquisition-costs/"),
        ("NA Mobile CPI $5.28", "BusinessOfApps UA Costs", "https://www.businessofapps.com/marketplace/user-acquisition/research/user-acquisition-costs/"),
        ("", "", ""),
        ("MARKET DATA", "", ""),
        ("Fantasy sports market $32B, 14.1% CAGR", "Mordor Intelligence", "https://www.mordorintelligence.com/industry-reports/north-america-fantasy-sports-market"),
        ("Sleeper $40M Series C, 3M+ users", "PR Newswire", "https://www.prnewswire.com/news-releases/fantasy-sports-app-sleeper-raises-40-million-doubles-user-base-in-2021-301379179.html"),
        ("", "", ""),
        ("METHODOLOGY NOTES", "", ""),
        ("User growth model", "Bottom-up monthly cohort model with seasonal multipliers", ""),
        ("Conversion timing", "In-season 1.8x multiplier, off-season 0.4x", ""),
        ("Churn model", "Monthly churn with seasonal variation (in vs off-season)", ""),
        ("Revenue recognition", "Subscription recognized monthly; annual billing amortized", ""),
        ("Marketing timing", "Digital marketing gated on launch month; influencer spend is lump-sum per month", ""),
        ("Platform fee allocation", "70% mobile (60/40 iOS/Android), 30% web (Stripe)", ""),
        ("COGS scaling", "Supabase upgrades at 50K users; Claude API scales per-query", ""),
        ("Advertising", "Web-only ads (no mobile app ads); 30% web × 80% no ad-blocker = 24% eligible", ""),
        ("DFS compliance (Canada)", "NO LICENSE REQUIRED — skill contest under Criminal Code s.206(1)(d); prudent compliance at scale", ""),
        ("DFS launch", "Month 19 (Oct 2027) per product roadmap", ""),
        ("", "", ""),
        ("US DFS EXPANSION (TOGGLE — SET TO 99 = OFF)", "", ""),
        ("NY Gaming Commission — $500K Y1", "Most expensive US state for DFS", "https://www.gaming.ny.gov/fantasysports"),
        ("State-by-state licensing fees", "19 states with DFS registration/licensing", "https://www.legalsportsreport.com/dfs-bill-tracker/"),
        ("GeoComply geolocation vendor", "Required for multi-state compliance", "https://www.geocomply.com"),
        ("NY DFS revenue tax ~15%", "Tax on DFS revenue from NY users", "https://www.gaming.ny.gov/fantasysports"),
        ("Estimated all-state Y1 cost ~$850K", "Licensing fees only; +$285K/yr compliance", ""),
        ("", "", ""),
        ("METHODOLOGY NOTES (continued)", "", ""),
        ("US DFS expansion toggle", "Set month index on Assumptions tab (99 = OFF); Y1 fees amortized /12, then annual renewal /12", ""),
        ("US DFS states modeled (19)", "NY, MA, VA, TN, IN, PA, CT, CO, IL, MO, OR, MD, KS, VT, AR, MS, NH, DE, RI", ""),
        ("US banned states", "Idaho, Hawaii banned; Nevada requires full gaming license (~$500K+)", ""),
        ("", "", ""),
        ("DYNAMIC MODEL NOTE", "", ""),
        ("All computed cells use Excel formulas", "Change any yellow input cell to recalculate instantly", ""),
        ("Input cells on Assumptions tab", "All model inputs: pricing, rates, costs, AND monthly new user targets", ""),
        ("Input cells on COGS tab", "Monthly conference/travel costs", ""),
    ]

    r = 2
    for dp, source, url in sources:
        ws9.cell(row=r, column=1, value=dp)
        if dp and dp.isupper():
            ws9.cell(row=r, column=1).font = section_font
        ws9.cell(row=r, column=2, value=source)
        if url:
            ws9.cell(row=r, column=3, value=url)
            ws9.cell(row=r, column=3).font = Font(color="0563C1", underline="single")
        r += 1

    # ── Freeze panes on all monthly tabs ──
    for ws_tab in [ws2, ws3, ws4, ws5, ws6]:
        ws_tab.freeze_panes = "B2"
    ws7.freeze_panes = "B2"

    # ── Save ──
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Citrus_Business_Model.xlsx")
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    path = build_workbook()
    print(f"\nDynamic workbook saved to: {path}")
    print("All computed cells use Excel formulas.")
    print("Edit any yellow input cell -> entire model recalculates instantly.")
    print("\nInput locations:")
    print("  - Assumptions tab: pricing, rates, costs, AND monthly new user targets (column B)")
    print("  - COGS tab: monthly conference costs (row 28)")
